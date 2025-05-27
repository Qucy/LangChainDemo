import csv
import re
import openpyxl
from openpyxl.styles import Font, Alignment

def convert_to_excel(input_file, output_file):
    # Initialize variables
    current_category = ""
    data_rows = []
    
    # Read the input file
    with open(input_file, 'r', encoding='utf-8') as file:
        lines = file.readlines()
    
    # Process each line
    for line in lines:
        line = line.strip()
        
        # Skip empty lines
        if not line:
            continue
        
        # Check if line is a category header
        if line.startswith('##'):
            current_category = line[2:].strip()  # Remove the ## prefix and any whitespace
            continue
        
        # Normalize the line - ensure all fields are properly quoted
        normalized_line = normalize_line(line)
        
        # Process data lines using regex to split by commas but not commas inside quotes
        parts = re.findall(r'"([^"]*)"', normalized_line)
        
        if len(parts) >= 4:
            capability = parts[0]
            customer_journey = parts[1]
            sample_question = parts[2]
            answer = parts[3]
            
            # Extract source URL from the answer
            source_url = extract_url(answer)
            
            # Remove the source URL from the answer if it exists
            clean_answer = remove_url_reference(answer)
            
            # Add to data rows
            data_rows.append([
                current_category,
                capability,
                customer_journey,
                sample_question,
                clean_answer,
                source_url
            ])
    
    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Converted Data"
    
    # Define headers
    headers = ['Category', 'Capability', 'Customer Journey', 'Sample Question', 'Answer', 'Source']
    
    # Write headers with formatting
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Write data rows
    for row_num, row_data in enumerate(data_rows, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width = min(max_length + 2, 50)  # Cap width at 50 characters
        ws.column_dimensions[column].width = adjusted_width
    
    # Save the workbook
    wb.save(output_file)
    
    return len(data_rows)

def extract_url(text):
    """
    Extract URL from text. Typically URLs are in the format (https://...)
    """
    url_match = re.search(r'https?://[^\s\)"]+', text)
    if url_match:
        return url_match.group(0)
    return ""

def remove_url_reference(text):
    """
    Remove URL references like (https://...) from the text
    """
    return re.sub(r'\s*\(https?://[^\)]+\)', '', text)

def normalize_line(line):
    """
    Ensure all fields in the line are properly quoted.
    This handles cases where fields don't start with quotes.
    """
    # If the line is already properly formatted with quotes, return it as is
    if re.match(r'^"[^"]*"(,"[^"]*")*$', line):
        return line
    
    # Split by commas, but preserve commas inside quotes
    parts = []
    in_quotes = False
    current_part = ""
    
    for char in line:
        if char == '"':
            in_quotes = not in_quotes
            current_part += char
        elif char == ',' and not in_quotes:
            # End of a part
            parts.append(current_part)
            current_part = ""
        else:
            current_part += char
    
    # Add the last part
    if current_part:
        parts.append(current_part)
    
    # Ensure each part is properly quoted
    normalized_parts = []
    for part in parts:
        part = part.strip()
        if not part.startswith('"'):
            part = '"' + part
        if not part.endswith('"'):
            part = part + '"'
        normalized_parts.append(part)
    
    return ','.join(normalized_parts)

if __name__ == "__main__":
    input_file = "PWS_QA.txt"
    output_file = "output.xlsx"  # Changed extension to .xlsx
    
    rows_processed = convert_to_excel(input_file, output_file)
    print(f"Conversion complete! {rows_processed} rows of data processed.")
    print(f"Excel file saved as: {output_file}")